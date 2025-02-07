
'''
import os
import csv
from openpyxl import load_workbook
from django.http import JsonResponse
from django.core.files.storage import FileSystemStorage
from dotenv import load_dotenv
from django.views.decorators.csrf import csrf_exempt
from .models import ProcessedFile


'''
# This is where the actual business logic for file processing, 
# interacting with the OpenAI API, and returning the processed data happens.
'''

# Load environment variables
load_dotenv()

# Your OpenAI API key setup (replace this with actual key management)
import openai
# print(os.getenv("OPENAI_API_KEY"))
openai.api_key = os.getenv("OPENAI_API_KEY")



@csrf_exempt
def process_file(request):
    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        user_input = request.POST.get('user_input')


        # Check if file is uploaded
        if not uploaded_file:
            return JsonResponse({'error': 'No file uploaded'}, status=400)

        # Check if user input is provided
        if not user_input:
            return JsonResponse({'error': 'No user input provided for pattern matching'}, status=400)

        # Handle file storage
        fs = FileSystemStorage()
        filename = fs.save(uploaded_file.name, uploaded_file)
        file_path = fs.url(filename)  # This will give you the path to the file

        # Check the file extension
        file_extension = filename.split('.')[-1].lower()

        rows = []

        try:
            # Parse CSV files
            if file_extension == 'csv':
                with fs.open(filename) as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row in reader:
                        rows.append(row)

            # Parse XLSX files
            elif file_extension in ['xlsx', 'xls']:
                workbook = load_workbook(fs.open(filename))
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1]]  # First row as headers
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))

                

            else:
                return JsonResponse({'error': 'Unsupported file format'}, status=400)

        except Exception as e:
            return JsonResponse({'error': f'Error reading file: {str(e)}'}, status=500)

        # Call OpenAI API to generate regex pattern
        try:
            regex_pattern, column_name, displacement = get_regex_from_openai(user_input)
        except Exception as e:
            return JsonResponse({'error': f'Error generating regex: {str(e)}'}, status=500)

        # Apply regex and replace values
        processed_data = apply_regex_to_data(rows, column_name, displacement)

        #delete the file after reading
        fs.delete(filename)


        # After processing data and before returning response
        processed_record = ProcessedFile.objects.create(
            filename=uploaded_file.name,
            user_input=user_input,
            processed_data={
                'data': processed_data,
                'regex_pattern': regex_pattern
            }
        )

        return JsonResponse({'processed_data': processed_data, 'regex_pattern': regex_pattern}, status=200)
    

    return JsonResponse({'error': 'Invalid request method'}, status=405)

def get_regex_from_openai(user_input):

    # replace the next line symbol with a space
    user_input = user_input.replace('\n', '')

    # Placeholder for calling OpenAI API to generate regex from user input
    # Replace this with your actual OpenAI call
    response = openai.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": f"Generate a regex pattern for: {user_input}, \
                   and only give me the regex_pattern, the column_name the user_input want to replace, \
                   and the dispacement user want to use like, only showing the info in ' ' 'regex_pattern:..., column_name:..., displacement:....' that's all do not add any other information"}],
        max_tokens=100,
        temperature=0,
    )
    result =  response.choices[0].message.content

    # Parse the response from OpenAI (Assuming the response is well-formed)
    # Extract regex pattern, column name, and displacement from the returned content
    
    x=1

    result = result.split(' ')
    

    regex_pattern = result[0].split('regex_pattern:')[1].strip()[:-1]
    column_name = result[1].split('column_name:')[1].strip()[:-1]
    displacement = result[2].split('displacement:')[1].strip()[:-1]
    
    return regex_pattern, column_name, displacement



def apply_regex_to_data(data, column_name, displacement):
    """
    This function checks if the value in the specified column matches the column_name (case-insensitively).
    If it matches, the entire column's data is replaced with the given displacement (replacement value).
    """
    processed_data = []
    

    for row in data:
        processed_row = {}
        for key, value in row.items():
            if key.lower() == column_name.lower():
                # If the value in the column matches the column name (case-insensitively),
                # replace the entire column data with displacement
                processed_row[key] = displacement
            else:
                processed_row[key] = value
        processed_data.append(processed_row)

    return processed_data

'''


import os
import csv
import logging
from openpyxl import load_workbook
from django.http import JsonResponse
from django.core.files.storage import FileSystemStorage
from dotenv import load_dotenv
from django.views.decorators.csrf import csrf_exempt
from .models import ProcessedFile
import openai

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Constants
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
ALLOWED_TYPES = [
    'text/csv',
    'application/vnd.ms-excel',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
]
openai.api_key = os.getenv("OPENAI_API_KEY")


def validate_file(uploaded_file):
    """Validate file size, type, and structure."""
    if uploaded_file.size > MAX_FILE_SIZE:
        raise ValueError("File size exceeds the maximum limit of 10MB.")
    
    if uploaded_file.content_type not in ALLOWED_TYPES:
        raise ValueError("Unsupported file type. Only CSV and Excel files are allowed.")
    
    # Validate CSV header
    if uploaded_file.content_type == 'text/csv':
        with uploaded_file.open('r') as f:
            header = f.readline().decode('utf-8').strip().split(',')
            if not header:
                raise ValueError("Invalid CSV file: No header found.")


def parse_csv(file_path):
    """Parse a CSV file and return rows as a list of dictionaries."""
    rows = []
    with open(file_path, mode='r', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            rows.append(row)
    return rows


def parse_excel(file_path):
    """Parse an Excel file and return rows as a list of dictionaries."""
    rows = []
    workbook = load_workbook(file_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]  # First row as headers
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


def get_regex_from_openai(user_input):

    # replace the next line symbol with a space
    user_input = user_input.replace('\n', '')
    
    """Call OpenAI API to generate regex pattern, column name, and displacement."""
    try:
        response = openai.chat.completions.create(
            model="gpt-4o",
            messages=[{
                "role": "user",
                "content": f"Generate a regex pattern for: {user_input}, "
                           f"and only give me the regex_pattern, the column_name the user_input wants to replace, "
                           f"and the displacement user wants to use. Format: 'regex_pattern:..., column_name:..., displacement:...'"
            }],
            max_tokens=100,
            temperature=0,
        )
        result = response.choices[0].message.content

        result = result.split(' ')

        regex_pattern = result[0].split('regex_pattern:')[1].strip()[:-1]
        column_name = result[1].split('column_name:')[1].strip()[:-1]
        displacement = result[2].split('displacement:')[1].strip()[:-1]
        
        return regex_pattern, column_name, displacement
    

    except Exception as e:
        logger.error(f"Error generating regex from OpenAI: {str(e)}")
        raise ValueError("Failed to generate regex pattern from OpenAI.")


def apply_regex_to_data(data, column_name, displacement):
    """Apply regex replacement to the specified column in the data."""
    processed_data = []
    for row in data:
        processed_row = {}
        for key, value in row.items():
            if key.lower() == column_name.lower():
                processed_row[key] = displacement
            else:
                processed_row[key] = value
        processed_data.append(processed_row)
    return processed_data


@csrf_exempt
def process_file(request):
    """Handle file upload, processing, and response."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)

    try:
        uploaded_file = request.FILES.get('file')
        user_input = request.POST.get('user_input')

        # Validate inputs
        if not uploaded_file:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
        if not user_input:
            return JsonResponse({'error': 'No user input provided'}, status=400)

        # Validate file
        validate_file(uploaded_file)

        # Save the file temporarily
        fs = FileSystemStorage()
        filename = fs.save(uploaded_file.name, uploaded_file)
        file_path = fs.path(filename)

        # Parse the file
        file_extension = filename.split('.')[-1].lower()
        if file_extension == 'csv':
            rows = parse_csv(file_path)
        elif file_extension in ['xlsx', 'xls']:
            rows = parse_excel(file_path)
        else:
            return JsonResponse({'error': 'Unsupported file format'}, status=400)

        # Generate regex and process data
        regex_pattern, column_name, displacement = get_regex_from_openai(user_input)
        processed_data = apply_regex_to_data(rows, column_name, displacement)

        # Save processed data to the database
        processed_record = ProcessedFile.objects.create(
            filename=uploaded_file.name,
            user_input=user_input,
            processed_data={
                'data': processed_data,
                'regex_pattern': regex_pattern
            }
        )

        # Clean up
        fs.delete(filename)

        return JsonResponse({
            'processed_data': processed_data,
            'regex_pattern': regex_pattern
        }, status=200)

    except ValueError as e:
        logger.error(f"Validation error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        return JsonResponse({'error': 'An unexpected error occurred'}, status=500)